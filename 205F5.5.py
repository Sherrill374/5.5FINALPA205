# FinalExam.py
# This program collects numeric and income data, appends it to final.csv,
# and creates both Excel and matplotlib charts.

import csv
import datetime
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
#1️askUser(): loop to ask for five numbers and display their total
def askUser():
    total = 0
    for i in range(5):  # Loop runs 5 times
        num = int(input("Please enter a number: "))
        total += num
    print(f"The sum for the 5 numbers entered is: {total}")

# 2️ askIncome(): loop to collect names and incomes and append to final.csv
def askIncome():
    with open("final.csv", "a", newline="") as file:
        writer = csv.writer(file)
        for i in range(5):  # Loop runs 5 times
            name = input("Please enter a name: ")
            income = int(input("Please enter their income: "))
            writer.writerow([name, income])

# 3️ creat excel pie chart from final.csv
def excelPie():
    wb = Workbook()
    ws = wb.active
    ws.title = "Income Data"

    # 1. Read from your CSV and write to the Excel sheet
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            # Converts the income string to an integer so Excel can graph it
            if row: 
                try:
                    ws.append([row[0], int(row[1])])
                except ValueError:
                    ws.append(row) # Keeps text headers intact if present

    # 2. Set up the pie chart
    pie = PieChart()
    
    # Data is column 2 (Income), categories/labels are column 1 (Names)
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    labels = Reference(ws, min_col=1, min_row=1, max_row=ws.max_row)
    
    pie.add_data(data, titles_from_data=False)
    pie.set_categories(labels)
    
    # Title with Student ID and date
    today = datetime.date.today().strftime("%m %d, %Y")
    pie.title = f"Lebshe2207 {today}"
    
    # 3. Add to sheet and save
    ws.add_chart(pie, "D2") # Placed next to data columns so it doesn't overlap
    wb.save("final.xlsx")

# 4️verticalBar(): create matplotlib bar chart from final.csv
def verticalBar():
    names = []
    incomes = []
    with open("final.csv", "r") as file:
        reader = csv.reader(file)
        for row in reader:
            names.append(row[0])
            incomes.append(int(row[1]))

    today = datetime.date.today().strftime("%m %d, %Y")
    plt.bar(names, incomes, color="g", label="Income")
    plt.title(f"Lebshe2207 {today}")
    plt.xlabel("Name")
    plt.ylabel("Income")
    plt.legend()
    plt.show()

# Run all functions
askUser()
askIncome()
excelPie()
verticalBar()
